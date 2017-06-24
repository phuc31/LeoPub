import xlrd
import openpyxl
import os
# All parameter
#workingDir      = 'D:\Workspace\PhucGiang\ReleaseJune/newPCL\PCL_test'
workingDir      = 'N:/prj_MBD/02_projects/03_management/development_process/16S-17K/06_Test_Preparation'
# format directory
workingDir      = workingDir.replace('\\', '/')
#WBOOKNAME    = 'ECPILS_V4.02.00_PCL_IT_R04 .xlsx'
notDataSheets  = ['Preface', 'Test Rule',
                'Test Result Outline',
                'Issue List', 'Revision History']
# Class save data 
class CData(object):
    normal      = 0
    abnormal    = 0
    boundary    = 0
    total       = 0

resultWBook = openpyxl.Workbook()
resultSheet = resultWBook['Sheet']
resultRow   = 1
resultCol   = 1

    # main function
def main():
    global resultWBook
    global resultSheet
    allFileData = CData()
    # Prevent ~file auto save of excel
    # os list dir get workbook name
    
    scriptDir = os.getcwd()
    os.chdir(workingDir)
    
    write_header_of_final_book()
    
    wBookNameList = os.listdir(workingDir)
    for wBookName in wBookNameList:
        if wBookName.find('_UT_') != -1 or wBookName.find('_IT_') != -1:
            bookData = collect_data_workbook(wBookName)
        
            write_then_tab('Sum of Book:')
            # write data of each sheet
            write_then_tab(bookData.total)
            write_then_tab(bookData.normal)
            write_then_tab(bookData.abnormal)
            write_then_tab(bookData.boundary)
        
            accumulate_data(bookData, allFileData)
            return_to_next_row(1)
        
    return_to_next_row(1)
    write_then_tab('Sum of All Book:')
    # write data of each sheet
    write_then_tab('')
    write_then_tab(allFileData.total)
    write_then_tab(allFileData.normal)
    write_then_tab(allFileData.abnormal)
    write_then_tab(allFileData.boundary)
    # analysis data
    os.chdir(scriptDir)
    # Save to file
    resultWBook.save('final_result.xlsx')
        
def collect_data_workbook(wBookName):
    global resultSheet
    global resultRow 
    global resultCol
    bookData        = CData()
    
    print '*** Loading Workbook: ' + wBookName
    wBook = xlrd.open_workbook(wBookName, encoding_override = 'cp932')
    print '*** Loaded'
    
    # Write work book name to result excel file then enter
    resultSheet.cell(row = resultRow,column = resultCol).value = wBookName
    return_to_next_row(2)
    
    # Scan all worksheet except not data worksheet
    sheetNameList = wBook.sheet_names()
    for sheetIndex in xrange(len(sheetNameList)):
        currentSheetName = sheetNameList[sheetIndex]
        print '***--- Worksheet: ' + currentSheetName
        wSheet = wBook.sheet_by_index(sheetIndex)
        
        # Scan sheets except unnecessary sheet
        if not currentSheetName in notDataSheets:
            # Get data of worksheet
            sheetData = collect_data_worksheet(wSheet)
            print 'Total: ' + str(sheetData.total)
            print 'Normal: ' + str(sheetData.normal)
            print 'Abnormal: '+ str(sheetData.abnormal)
            print 'Boundary' + str(sheetData.boundary)
            
            # calculate all data of a book
            bookData.total    = bookData.total + sheetData.total
            bookData.normal   = bookData.normal + sheetData.normal
            bookData.abnormal = bookData.abnormal + sheetData.abnormal
            bookData.boundary = bookData.boundary + sheetData.boundary
            
            # write sheet name
            write_then_tab(currentSheetName)
            
            # write data of each sheet
            write_then_tab(sheetData.total)
            write_then_tab(sheetData.normal)
            write_then_tab(sheetData.abnormal)
            write_then_tab(sheetData.boundary)
            
            return_to_next_row(2)
            # Done a sheet
            
    print 'Done'
    return bookData
    
def collect_data_worksheet(wSheet):
    sheetData = CData()
    # Search and get data:
    # Total, Normal, Abnormal, Boundary
    for rowIndex in range(wSheet.nrows):
        row = wSheet.row(rowIndex)
        i = 0
        for cell in row:
            if cell.value is None:
                # check whether cell is empty
                pass
            else:
                if cell.ctype == xlrd.XL_CELL_TEXT:
                    #print 'cell value: '
                    # print cell.value.encode('latin-1', 'ignore')
                    if cell.value.find('Normality') != -1:
                        if row[i + 1].ctype == xlrd.XL_CELL_NUMBER:
                            sheetData.normal = row[i + 1].value
                            # print row[i + 1].ctype
                            # print 'value: ' + str(row[i+1].value)
                    if cell.value.find('Abnormality') != -1:
                        if row[i + 1].ctype == xlrd.XL_CELL_NUMBER:
                            sheetData.abnormal = row[i + 1].value
                            # print row[i + 1].ctype
                            # print 'value: ' + str(row[i+1].value)
                    if cell.value.find('Boundary') != -1:
                        if row[i + 1].ctype == xlrd.XL_CELL_NUMBER:
                            sheetData.boundary = row[i + 1].value
                            # print row[i + 1].ctype
                            # print 'value: ' + str(row[i+1].value)
                    if cell.value.find('Total') != -1:
                        if row[i + 1].ctype == xlrd.XL_CELL_NUMBER:
                            # print row[i + 1].ctype
                            # print 'value: ' + str(row[i+1].value)
                            sheetData.total = row[i + 1].value
            i = i + 1
    return sheetData

## These function support write to final result  book
def write_then_tab(data):
    global resultRow
    global resultCol
    global resultSheet
    
    resultSheet.cell(row = resultRow, column = resultCol).value = data
    resultCol = resultCol + 1
    
def return_to_next_row(startCol):
    global resultRow
    global resultCol
    
    resultCol = startCol
    resultRow = resultRow + 1

def write_header_of_final_book():
    # Write Header
    write_then_tab('')
    write_then_tab('')
    write_then_tab('Total')
    write_then_tab('Normal')
    write_then_tab('Abnormal')
    write_then_tab('Boundary')
    return_to_next_row(1)

def accumulate_data(srcData, dstData):
    dstData.total = dstData.total + srcData.total
    dstData.normal = dstData.normal + srcData.normal
    dstData.abnormal = dstData.abnormal + srcData.abnormal
    dstData.boundary = dstData.boundary + srcData.boundary
    return dstData
    
#def analysis_data_of_all_file():
    

if __name__ == "__main__": main()